import openpyxl, datetime

# VARIABLE SECTION #
DATETIME = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

# Saving the data in an Excel file
workbook = openpyxl.load_workbook(filename = "lego.xlsx")
sheet = workbook['Price database']

old_items = [sheet[openpyxl.utils.cell.get_column_letter(x)][0].value for x in range(1, sheet.max_column, 3)]

for tupl in prices.items():
	
	row = 1

	# Checking wheter the item name already exists in the database
	if (tupl[0] in old_items):
		# If yes, write the new price after the last row for that item

		# Column
		column = ((old_items.index(tupl[0])) * 3) + 1
		#print(str(column) + "-" + openpyxl.utils.cell.get_column_letter(column))

		# Row
		for el in sheet[openpyxl.utils.cell.get_column_letter(column)]:
			if el.value != None:
					row += 1
			else: break
		
	else:
		# Otherwise, start writing right after the last item name
			column = sheet.max_column + 1 if sheet.max_column > 1 else 1
			sheet.cell(row = row, column = column).value = tupl[0]
			sheet.merge_cells(openpyxl.utils.cell.get_column_letter(column) + str(row) + ":" + openpyxl.utils.cell.get_column_letter(column + 2) + str(row))
			row += 1

	counter = 0
	for item in tupl[1]:
			sheet.cell(row = row + counter, column = column).value = DATETIME
			sheet.cell(row = row + counter, column = column + 1).value = float(item[0]) ; sheet.cell(row = row + counter, column = column + 1).number_format = '#,##0.00€'
			sheet.cell(row = row + counter, column = column + 2).value = item[1]
			counter += 1


workbook.save(filename  = "Lego.xlsx")